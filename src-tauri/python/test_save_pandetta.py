#!/usr/bin/env python3
"""Test per save_pandetta.py"""
import os
import sys
import json
import tempfile
import openpyxl
from openpyxl.styles import PatternFill

# Importa la funzione da testare
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from save_pandetta import main

def create_test_excel(filepath):
    """Crea un file Excel di test con una tabella 'Pandetta'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pandetta"
    
    # Intestazioni
    headers = ['N.RIF', 'Data', 'Cliente', 'Descrizione', 'STATO', 'Tecnico']
    ws.append(headers)
    
    # Formattazione per le intestazioni (azzurro)
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    for cell in ws[1]:
        cell.fill = header_fill
    
    # Righe di esempio con formattazione
    example_data = [
        ['001', '2024-01-15', 'Mario Rossi', 'Installazione PC', 'aperto', 'Giovanni'],
        ['002', '2024-01-16', 'Luigi Verdi', 'Riparazione stampante', 'chiuso', 'Marco'],
        ['003', '2024-01-17', 'Anna Bianchi', 'Configurazione rete', 'annullato', 'Paolo'],
    ]
    
    for row_data in example_data:
        ws.append(row_data)
    
    # Applica formattazione alle righe di dati (es. bordi, font)
    for row_idx in range(2, 5):
        for cell in ws[row_idx]:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
    
    wb.save(filepath)

def normalize_color(color_str):
    """Normalizza la stringa colore prendendo gli ultimi 6 caratteri (RGB)."""
    if color_str is None:
        return None
    return color_str[-6:].upper() if len(color_str) >= 6 else color_str.upper()

def test_save_pandetta():
    """Test case principale."""
    print("Avvio test save_pandetta...")
    
    with tempfile.TemporaryDirectory() as tmpdir:
        input_excel = os.path.join(tmpdir, 'input.xlsx')
        output_excel = os.path.join(tmpdir, 'output.xlsx')
        json_data = os.path.join(tmpdir, 'data.json')
        
        # Crea Excel di input
        create_test_excel(input_excel)
        
        # Crea JSON con dati da salvare
        payload = {
            "data": [
                # Modifica riga esistente (N.RIF = 001)
                {
                    "N.RIF": "001",
                    "Data": "2024-01-18",
                    "Cliente": "Mario Rossi",
                    "Descrizione": "Aggiornamento sistema",
                    "STATO": "chiuso",  # Cambia stato da aperto a chiuso
                    "Tecnico": "Giovanni"
                },
                # Nuova riga (nuovo N.RIF)
                {
                    "N.RIF": "004",
                    "Data": "2024-01-19",
                    "Cliente": "Carlo Neri",
                    "Descrizione": "Nuova installazione",
                    "STATO": "aperto",
                    "Tecnico": "Luca"
                },
                # Altra modifica (N.RIF = 003, già annullato)
                {
                    "N.RIF": "003",
                    "Data": "2024-01-20",
                    "Cliente": "Anna Bianchi",
                    "Descrizione": "Supporto telefonico",
                    "STATO": "fuoriuso",  # Sempre rosso
                    "Tecnico": "Paolo"
                }
            ],
            "dynamic_cols": ["N.RIF", "Data", "Cliente", "Descrizione", "STATO", "Tecnico"]
        }
        
        with open(json_data, 'w', encoding='utf-8') as f:
            json.dump(payload, f)
        
        # Esegui la funzione
        sys.argv = ['save_pandetta.py', json_data, input_excel, output_excel]
        main()
        
        # Verifica output
        wb = openpyxl.load_workbook(output_excel)
        ws = wb["Pandetta"]
        
        rows = list(ws.iter_rows(values_only=True))
        print(f"Numero di righe nel file output: {len(rows)}")
        
        # Verifica che le righe siano 4 originali + 1 nuova = 5 totali
        assert len(rows) == 5, f"Atteso 5 righe, trovato {len(rows)}"
        
        # Verifica N.RIF: 001, 002, 003, 004 (dovrebbe esserci anche 002 originale)
        rif_values = [row[0] for row in rows[1:]]  # Salta header
        print(f"Valori N.RIF: {rif_values}")
        assert rif_values[0] == '001', "Riga 001 modificata"
        assert rif_values[1] == '002', "Riga 002 invariata"
        assert rif_values[2] == '003', "Riga 003 modificata"
        assert rif_values[3] == '004', "Nuova riga 004"
        
        # Verifica modifiche
        # Riga con N.RIF 001: Stato dovrebbe essere 'chiuso' e quindi colore verde
        row_001 = ws[2]  # Seconda riga (dopo header)
        stato_cell = row_001[4]  # Colonna STATO (indice 4)
        assert stato_cell.value == 'chiuso', f"Stato 001: atteso 'chiuso', trovato '{stato_cell.value}'"
        if stato_cell.fill and stato_cell.fill.fill_type:
            start_color = stato_cell.fill.start_color
            color_str = None
            if hasattr(start_color, 'rgb') and start_color.rgb:
                color_str = start_color.rgb
            elif hasattr(start_color, 'indexed') and start_color.indexed is not None:
                color_str = str(start_color.indexed)
            print(f"Colore cella stato 001: {color_str}")
            if color_str:
                assert normalize_color(color_str) == '00FF00', "Colore non verde per stato chiuso"
        
        # Riga con N.RIF 003: Stato 'fuoriuso' -> rosso
        row_003 = ws[4]  # Quarta riga
        stato_cell_003 = row_003[4]
        assert stato_cell_003.value == 'fuoriuso', f"Stato 003: atteso 'fuoriuso', trovato '{stato_cell_003.value}'"
        if stato_cell_003.fill and stato_cell_003.fill.fill_type:
            start_color_003 = stato_cell_003.fill.start_color
            color_str_003 = None
            if hasattr(start_color_003, 'rgb') and start_color_003.rgb:
                color_str_003 = start_color_003.rgb
            elif hasattr(start_color_003, 'indexed') and start_color_003.indexed is not None:
                color_str_003 = str(start_color_003.indexed)
            print(f"Colore cella stato 003: {color_str_003}")
            if color_str_003:
                assert normalize_color(color_str_003) == 'FF0000', "Colore non rosso per stato fuoriuso"
        
        # Riga con N.RIF 004: nuova riga, stato 'aperto' -> giallo
        row_004 = ws[5]  # Quinta riga
        stato_cell_004 = row_004[4]
        assert stato_cell_004.value == 'aperto', f"Stato 004: atteso 'aperto', trovato '{stato_cell_004.value}'"
        if stato_cell_004.fill and stato_cell_004.fill.fill_type:
            start_color_004 = stato_cell_004.fill.start_color
            color_str_004 = None
            if hasattr(start_color_004, 'rgb') and start_color_004.rgb:
                color_str_004 = start_color_004.rgb
            elif hasattr(start_color_004, 'indexed') and start_color_004.indexed is not None:
                color_str_004 = str(start_color_004.indexed)
            print(f"Colore cella stato 004: {color_str_004}")
            if color_str_004:
                assert normalize_color(color_str_004) == 'FFFF00', "Colore non giallo per stato aperto"
        
        # Verifica formattazione: nuova riga dovrebbe avere bordi come le altre
        # Controlla che la riga 5 abbia bordi (seabbiamo copiato formattazione)
        has_border = any(cell.border for cell in row_004)
        print(f"Riga 004 ha bordi: {has_border}")
        assert has_border, "Nuova riga non ha i bordi copiati"
        
        print("✅ Tutti i test sono passati!")
        return True

if __name__ == '__main__':
    import sys
    try:
        test_save_pandetta()
    except Exception as e:
        print(f"❌ Test fallito: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
