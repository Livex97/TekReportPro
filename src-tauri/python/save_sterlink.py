#!/usr/bin/env python3
import sys
import json
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string

def get_column_index(headers, column_name):
    """Trova l'indice della colonna con nome specifico (case-insensitive, gestisce varianti)."""
    column_name_clean = column_name.strip().upper().replace('.', '').replace(' ', '').replace('\n', '')
    
    for idx, h in enumerate(headers):
        if h:
            header_clean = str(h).strip().upper().replace('.', '').replace(' ', '').replace('\n', '')
            if header_clean == column_name_clean:
                return idx + 1
            if column_name_clean in header_clean or header_clean in column_name_clean:
                return idx + 1
    return None

def copy_row_formatting(source_row, target_row, ws):
    """Copia la formattazione da una riga sorgente a una riga destinazione."""
    for source_cell, target_cell in zip(source_row, target_row):
        if source_cell.has_style:
            target_cell.font = source_cell.font.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.alignment = source_cell.alignment.copy()
            target_cell.number_format = source_cell.number_format
            target_cell.protection = source_cell.protection.copy()
            if source_cell.fill and source_cell.fill.fill_type:
                target_cell.fill = source_cell.fill.copy()

def main():
    if len(sys.argv) != 4:
        print("Usage: save_sterlink.py <json> <input.xlsx> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    json_path, in_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]

    with open(json_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)
    
    current_data = payload.get('current_data', [])
    dynamic_cols = payload.get('dynamic_cols')
    original_rows_count = payload.get('original_rows_count')

    try:
        wb = openpyxl.load_workbook(in_path)
    except Exception as e:
        print(f"Error loading workbook: {e}", file=sys.stderr)
        sys.exit(1)
    
    ws = wb.active # Sterlink usually has one main sheet
    
    # Find the table in the worksheet
    table = None
    if hasattr(ws, 'tables') and ws.tables:
        # Look for the specific table name "Tabella1" first
        if 'Tabella1' in ws.tables:
            table = ws.tables['Tabella1']
        else:
            # Fallback to the first table if "Tabella1" not found
            table_name = list(ws.tables.keys())[0]
            table = ws.tables[table_name]

    if not dynamic_cols:
        if current_data:
            dynamic_cols = [k for k in current_data[0].keys() if not k.startswith('_')]
        else:
            raise ValueError("No columns provided")

    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h else '' for h in header_cells]
    col_map = {h: idx+1 for idx, h in enumerate(headers) if h}
    
    # Determine table boundaries if a table exists
    table_bounds = None
    if table:
        # Parse the table range (e.g., "A1:D10")
        try:
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            table_bounds = (min_row, max_row, min_col, max_col)
        except Exception:
            # If we can't parse the table ref, fall back to using the whole sheet
            table_bounds = None

    # ID colonna per Sterlink: NUMERO CHECKLIST
    id_col_idx = get_column_index(headers, 'NUMERO CHECKLIST')
    if id_col_idx is None:
        # Fallback to SERIALE if NUMERO CHECKLIST not found, but user asked for NUMERO CHECKLIST
        id_col_idx = get_column_index(headers, 'SERIALE')
        if id_col_idx is None:
            raise ValueError("Colonna 'NUMERO CHECKLIST' o 'SERIALE' non trovata nel file")

    # Mappa dei valori ID esistenti -> numero riga nel foglio
    existing_keys = {}
    example_row = None
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if id_col_idx <= len(row):
            key_cell = row[id_col_idx-1]
            if key_cell.value is not None:
                key_val = key_cell.value
                # Normalize numeric IDs
                if isinstance(key_val, float) and key_val.is_integer():
                    key_val = int(key_val)
                key_str = str(key_val).strip()
                if key_str:
                    existing_keys[key_str] = r_idx
                    if example_row is None:
                        example_row = row

    # Trova il nome esatto della colonna ID nelle dynamic_cols
    id_col_name = next((col for col in dynamic_cols if 'NUMERO' in col.upper() and 'CHECKLIST' in col.upper()), None)
    if not id_col_name:
        id_col_name = next((col for col in dynamic_cols if 'SERIALE' in col.upper()), None)

    # Raccogli ID delle righe correnti per individuare eliminazioni
    current_ids = set()
    for row in current_data:
        if row.get('_empty'):
            continue
        id_val = row.get(id_col_name) if id_col_name else None
        if id_val is not None:
            if isinstance(id_val, float) and id_val.is_integer():
                id_val = int(id_val)
            current_ids.add(str(id_val).strip())

    # Identifica righe da eliminare
    rows_to_delete = []
    for id_str, r_idx in existing_keys.items():
        if id_str not in current_ids:
            rows_to_delete.append(r_idx)

    # Processa le righe correnti
    for row in current_data:
        if row.get('_empty'):
            continue

        id_val = row.get(id_col_name) if id_col_name else None
        if isinstance(id_val, float) and id_val.is_integer():
            id_val = int(id_val)
        
        target_row = None
        key_str = str(id_val).strip() if id_val is not None else None
        
        if key_str and key_str in existing_keys:
            r = existing_keys[key_str]
            for col in dynamic_cols:
                if col in col_map:
                    ws.cell(row=r, column=col_map[col], value=row.get(col))
            target_row = ws[r]
        else:
            # Nuova riga - inseriscila correttamente nella tabella se esiste, altrimenti alla fine del foglio
            if table_bounds:
                # Inserisci la riga appena prima dell'ultima riga della tabella per mantenerla dentro i limiti della tabella
                min_row, max_row, min_col, max_col = table_bounds
                r = max_row + 1  # Aggiungi alla fine della tabella corrente
                
                # Inserisci la riga nel foglio
                ws.insert_rows(r)
                
                # Aggiorna il riferimento della tabella per includere la nuova riga
                # Il nuovo riferimento sarà dalla stessa colonna di inizio alla stessa colonna di fine,
                # ma dalla stessa riga di inizio alla nuova riga di fine
                new_ref = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=r, column=max_col).coordinate}"
                table.ref = new_ref
            else:
                # Nessuna tabella trovata, aggiungi alla fine del foglio (comportamento originale)
                r = ws.max_row + 1
            
            # Compila i valori delle colonne
            for col in dynamic_cols:
                if col in col_map:
                    ws.cell(row=r, column=col_map[col], value=row.get(col))
            
            target_row = ws[r]
            if example_row:
                copy_row_formatting(example_row, target_row, ws)

    # Elimina le righe rimosse
    rows_to_delete.sort(reverse=True)
    for r_idx in rows_to_delete:
        try:
            ws.delete_rows(r_idx)
        except Exception as e:
            print(f"Warning: could not delete row {r_idx}: {e}", file=sys.stderr)

    # Aggiorna la dimensione della tabella se presente
    if table:
        table_range = table.ref
        if ':' in table_range:
            start_cell_str, end_cell_str = table_range.split(':')
            
            # Estrai colonna iniziale e riga iniziale
            start_col_letters = ''.join([c for c in start_cell_str if not c.isdigit()])
            start_row = int(''.join([c for c in start_cell_str if c.isdigit()])) if any(c.isdigit() for c in start_cell_str) else 1
            
            # Estrai colonna finale
            end_col_letters = ''.join([c for c in end_cell_str if not c.isdigit()])
            
            # Trova indici numerici
            try:
                start_col_idx = column_index_from_string(start_col_letters)
                end_col_idx = column_index_from_string(end_col_letters)
            except Exception:
                start_col_idx = 1
                end_col_idx = ws.max_column

            # Trova l'ultima riga con dati
            last_row_with_data = start_row
            for row_idx in range(ws.max_row, start_row, -1):
                row_has_data = False
                for col_idx in range(start_col_idx, end_col_idx + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        last_row_with_data = row_idx
                        row_has_data = True
                        break
                if row_has_data:
                    break
            
            new_last_row = last_row_with_data
            new_end_cell = f"{end_col_letters}{new_last_row}"
            table.ref = f"{start_cell_str}:{new_end_cell}"
            if hasattr(table, 'autoFilter') and table.autoFilter:
                table.autoFilter.ref = table.ref

    # Riordina le righe dati per NUMERO CHECKLIST (crescente) prima di salvare
    # Questo risolve il problema delle nuove righe aggiunte in fondo al file
    rows_with_data = []
    for r_idx in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() for v in row_vals):
            id_val = ws.cell(row=r_idx, column=id_col_idx).value
            if isinstance(id_val, float) and id_val.is_integer():
                id_val = int(id_val)
            rows_with_data.append((id_val, r_idx, row_vals))
    
    # Ordina per ID (crescente), mettendo le righe senza ID in fondo
    def sort_key(item):
        id_val = item[0]
        if id_val is None or str(id_val).strip() == '':
            return (1, 0)  # Metti righe senza ID in fondo
        try:
            return (0, int(id_val))
        except (ValueError, TypeError):
            return (0, 0)
    
    rows_with_data.sort(key=sort_key)
    
    # Leggi le intestazioni per mantenerle
    header_vals = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    
    # Copia i dati riordinati
    for target_r_idx, (_, source_r_idx, row_vals) in enumerate(rows_with_data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=target_r_idx, column=c_idx).value = val
    
    # Pulisci righe vuote in fondo
    for r_idx in range(len(rows_with_data) + 2, ws.max_row + 2):
        for c_idx in range(1, ws.max_column + 1):
            ws.cell(row=r_idx, column=c_idx).value = None

    wb.save(out_path)
    print(f"Successfully saved to {out_path}")

if __name__ == '__main__':
    main()
