#!/usr/bin/env python3
import sys
import json
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string

STATUS_COLORS = {
    'aperta': PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),  # Giallo
    'chiusa': PatternFill(start_color='FF00B050', end_color='FF00B050', fill_type='solid'),  # Verde
    'negativa': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),  # Rosso
}

TECNICO_COLORS = {
    'MEZZAPESA': {
        'fill': PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid'),  # Grigio
        'font': Font(color='FFFFFFFF')  # Bianco
    },
    'ALLEGREZZA': {
        'fill': PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid'),  # Bianco
        'font': Font(color='FF000000')  # Nero
    },
    'AMARA': {
        'fill': PatternFill(start_color='FF4287F5', end_color='FF4287F5', fill_type='solid'),  # Blu
        'font': Font(color='FFFFFFFF')  # Bianco
    }
}

def get_column_index(headers, column_name):
    """Trova l'indice della colonna con nome specifico (case-insensitive, gestisce varianti)."""
    column_name_clean = column_name.strip().upper().replace('.', '').replace(' ', '')
    
    for idx, h in enumerate(headers):
        if h:
            header_clean = str(h).strip().upper().replace('.', '').replace(' ', '')
            if header_clean == column_name_clean:
                return idx + 1
            if column_name_clean in header_clean or header_clean in column_name_clean:
                return idx + 1
    return None

def copy_row_formatting(source_row, target_row, ws):
    """Copia la formattazione da una riga sorgente a una riga destinazione."""
    for source_cell, target_cell in zip(source_row, target_row):
        if source_cell.has_style:
            target_cell.font = source_cell.font.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.alignment = source_cell.alignment.copy()
            target_cell.number_format = source_cell.number_format
            target_cell.protection = source_cell.protection.copy()
            if source_cell.fill and source_cell.fill.fill_type:
                target_cell.fill = source_cell.fill.copy()

def derive_status(stato_val, esito_val):
    """Calcola lo stato in base a STATO INTERVENTO e ESITO."""
    stato = str(stato_val).strip().upper() if stato_val is not None else ""
    esito = str(esito_val).strip().upper() if esito_val is not None else ""

    if ('CHIUSO' in stato or 'CHIUSA' in stato) and 'POSITIVO' in esito:
        return 'chiusa'
    # Se Esito contiene NEGATIVO → negativa (rosso)
    if 'NEGATIVO' in esito:
        return 'negativa'
    if 'ANNULLAT' in stato or 'FUORI USO' in stato or 'NON RIPARABILE' in stato or 'IRREPARABILE' in stato or 'ANNULLAT' in esito or 'FUORI USO' in esito:
        return 'negativa'
    return 'aperta'

def apply_status_color(row_cells, status_value, ws):
    """Applica il colore di sfondo basato sullo Stato a tutta la riga."""
    if not status_value:
        return
    
    status_str = str(status_value).strip().lower()
    fill = STATUS_COLORS.get(status_str)
    
    if fill:
        for cell in row_cells:
            cell.fill = fill

def apply_tecnico_color(cell, tecnico_value, color_map_payload=None):
    """Applica colore specifico per valori TECNICO nella cella."""
    if not tecnico_value:
        return
    tecnico_str = str(tecnico_value).strip().upper()
    
    # 1. Prova prima con la mappa dal payload (più aggiornata)
    if color_map_payload and tecnico_str in color_map_payload:
        info = color_map_payload[tecnico_str]
        # Il frontend manda {bg, text, export} dove export è hex RRGGBB
        hex_color = info.get('export', 'FFFFFF').replace('#', '')
        if len(hex_color) == 6:
            hex_color = 'FF' + hex_color # Aggiungi alpha
        
        # Determina il colore del testo (luminosità)
        text_color = info.get('text', '#000000').replace('#', '')
        if len(text_color) == 6:
            text_color = 'FF' + text_color
            
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        cell.font = Font(color=text_color)
        return

    # 2. Fallback alla mappa hardcoded
    color_info = TECNICO_COLORS.get(tecnico_str)
    if color_info:
        cell.fill = color_info['fill']
        cell.font = color_info['font']

def main():
    if len(sys.argv) != 4:
        print("Usage: save_pandetta.py <json> <input.xlsx> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    json_path, in_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]

    with open(json_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)
    
    current_data = payload.get('current_data', [])
    original_data = payload.get('original_data', [])
    dynamic_cols = payload.get('dynamic_cols')
    original_rows_count = payload.get('original_rows_count')
    # Prendi la mappa dei colori dal payload se esiste
    tecnico_color_map_payload = payload.get('tecnico_color_map', {})

    wb = openpyxl.load_workbook(in_path)
    
    # Cerca specificamente la tabella 'Tabella1' in tutti i fogli
    pandetta_table = None
    ws = None
    for sheet in wb.worksheets:
        tables_list = getattr(sheet, 'tables', getattr(sheet, '_tables', {}))
        # Cerca prima la tabella con nome esatto "Tabella1"
        if 'Tabella1' in tables_list:
            pandetta_table = tables_list['Tabella1']
            ws = sheet
            break
        # Se non troviamo Tabella1, cerchiamo comunque una tabella Pandetta per compatibilità
        elif pandetta_table is None:  # Solo se non abbiamo già trovato Tabella1
            for table_name, table in tables_list.items():
                table_name_upper = table_name.upper() if not hasattr(table, 'name') else table.name.upper()
                if any(keyword in table_name_upper for keyword in ['PANDETTA', 'PANDET', 'ASSISTENZA']):
                    pandetta_table = table
                    ws = sheet
                    break
    
    # Se non troviamo nessuna tabella specifica, usiamo il foglio attivo
    if ws is None:
        ws = wb.active
    
    # Determina i limiti della tabella se esiste
    table_bounds = None
    if pandetta_table:
        try:
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(pandetta_table.ref)
            table_bounds = (min_row, max_row, min_col, max_col)
        except Exception:
            # Se non possiamo analizzare il riferimento della tabella, torniamo al comportamento originale
            table_bounds = None

    if not dynamic_cols:
        if current_data:
            dynamic_cols = [k for k in current_data[0].keys() if not k.startswith('_')]
        else:
            raise ValueError("No columns provided")

    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h else '' for h in header_cells]
    col_map = {h: idx+1 for idx, h in enumerate(headers) if h}

    rif_col_idx = get_column_index(headers, 'N.RIF')
    if rif_col_idx is None:
        raise ValueError("Colonna 'N.RIF' non trovata nel file")

    status_col_idx = get_column_index(headers, 'STATO')
    status_col_name = headers[status_col_idx-1] if status_col_idx is not None else None

    tecnico_col_idx = get_column_index(headers, 'TECNICO')

    # Mappa dei valori RIF esistenti -> numero riga nel foglio
    existing_keys = {}
    example_row = None  # Per copiare formattazione (prima riga di dati)
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if rif_col_idx <= len(row):
            key_cell = row[rif_col_idx-1]
            if key_cell.value is not None:
                key_val = key_cell.value
                if isinstance(key_val, float) and key_val.is_integer():
                    key_val = int(key_val)
                key_str = str(key_val).strip()
                if key_str:
                    existing_keys[key_str] = r_idx
                    if example_row is None:
                        example_row = row

    # Trova i nomi esatti delle colonne dinamiche
    dynamic_cols_upper = {col.upper(): col for col in dynamic_cols}
    rif_col_name = next((dynamic_cols_upper[k] for k in dynamic_cols_upper if 'RIF' in k), None)
    stato_col_name = next((dynamic_cols_upper[k] for k in dynamic_cols_upper if 'STATO' in k and 'INTERVENTO' in k), None)
    if not stato_col_name:
        stato_col_name = next((dynamic_cols_upper[k] for k in dynamic_cols_upper if 'STATO' in k), None)
    esito_col_name = next((dynamic_cols_upper[k] for k in dynamic_cols_upper if 'ESITO' in k), None)

    if original_rows_count is None:
        original_rows_count = len(current_data)

    # Raccogli RIF delle righe correnti per individuare eliminazioni
    current_rifs = set()
    for row in current_data:
        if row.get('_empty'):
            continue
        rif_val = row.get(rif_col_name) if rif_col_name else None
        if rif_val is not None:
            if isinstance(rif_val, float) and rif_val.is_integer():
                rif_val = int(rif_val)
            current_rifs.add(str(rif_val).strip())

    # Identifica righe da eliminare (presenti in originale ma non in corrente)
    rows_to_delete = []
    for rif_str, r_idx in existing_keys.items():
        if rif_str not in current_rifs:
            rows_to_delete.append(r_idx)

    # Processa le righe correnti (aggiorna esistenti o aggiunge nuove)
    for i, row in enumerate(current_data):
        if row.get('_empty'):
            continue

        rif_val = row.get(rif_col_name) if rif_col_name else None
        if isinstance(rif_val, float) and rif_val.is_integer():
            rif_val = int(rif_val)
        
        # Calcolo stato automatico
        stato_val = row.get(stato_col_name) if stato_col_name else None
        esito_val = row.get(esito_col_name) if esito_col_name else None
        status_value = derive_status(stato_val, esito_val)
        
        target_row = None
        
        if rif_val is None:
            # Se non c'è N.RIF, aggiungi nuova riga
            if table_bounds:
                # Inserisci la riga dentro i limiti della tabella
                min_row, max_row, min_col, max_col = table_bounds
                r = max_row + 1  # Aggiungi alla fine della tabella corrente
                
                # Inserisci la riga nel foglio
                ws.insert_rows(r)
                
                # Aggiorna il riferimento della tabella per includere la nuova riga
                # Il nuovo riferimento sarà dalla stessa colonna di inizio alla stessa colonna di fine,
                # ma dalla stessa riga di inizio alla nuova riga di fine
                new_ref = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=r, column=max_col).coordinate}"
                pandetta_table.ref = new_ref
                if hasattr(pandetta_table, 'autoFilter') and pandetta_table.autoFilter:
                    pandetta_table.autoFilter.ref = pandetta_table.ref
            else:
                # Nessuna tabella trovata o limiti non determinabili, aggiungi alla fine del foglio
                r = ws.max_row + 1
            
            for col in dynamic_cols:
                if col in col_map:
                    ws.cell(row=r, column=col_map[col], value=row.get(col))
            target_row = ws[r]
            if example_row:
                copy_row_formatting(example_row, target_row, ws)
        else:
            key_str = str(rif_val).strip()
            if key_str in existing_keys:
                r = existing_keys[key_str]
                for col in dynamic_cols:
                    if col in col_map:
                        ws.cell(row=r, column=col_map[col], value=row.get(col))
                target_row = ws[r]
            else:
                # Nuova rif - aggiungi nuova riga
                if table_bounds:
                    # Inserisci la riga dentro i limiti della tabella
                    min_row, max_row, min_col, max_col = table_bounds
                    r = max_row + 1  # Aggiungi alla fine della tabella corrente
                    
                    # Inserisci la riga nel foglio
                    ws.insert_rows(r)
                    
                    # Aggiorna il riferimento della tabella per includere la nuova riga
                    # Il nuovo riferimento sarà dalla stessa colonna di inizio alla stessa colonna di fine,
                    # ma dalla stessa riga di inizio alla nuova riga di fine
                    new_ref = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=r, column=max_col).coordinate}"
                    pandetta_table.ref = new_ref
                    if hasattr(pandetta_table, 'autoFilter') and pandetta_table.autoFilter:
                        pandetta_table.autoFilter.ref = pandetta_table.ref
                else:
                    # Nessuna tabella trovata o limiti non determinabili, aggiungi alla fine del foglio
                    r = ws.max_row + 1
                
                for col in dynamic_cols:
                    if col in col_map:
                        ws.cell(row=r, column=col_map[col], value=row.get(col))
                target_row = ws[r]
                if example_row:
                    copy_row_formatting(example_row, target_row, ws)
        
        if target_row and status_value:
            apply_status_color(target_row, status_value, ws)
        
        if target_row and tecnico_col_idx:
            if tecnico_col_idx <= len(target_row):
                tecnico_cell = target_row[tecnico_col_idx - 1]
                apply_tecnico_color(tecnico_cell, tecnico_cell.value, tecnico_color_map_payload)

    # Elimina le righe non più presenti (in ordine decrescente per evitare problemi di indice)
    rows_to_delete.sort(reverse=True)
    for r_idx in rows_to_delete:
        try:
            ws.delete_rows(r_idx)
        except Exception as e:
            print(f"Warning: could not delete row {r_idx}: {e}", file=sys.stderr)

    # Espandi la tabella per includere le nuove righe se necessario
    if pandetta_table:
        table_range = pandetta_table.ref
        if ':' in table_range:
            start_cell_str, end_cell_str = table_range.split(':')
            
            # Estrai colonna e riga iniziale
            start_col_letters = ''.join([c for c in start_cell_str if not c.isdigit()])
            start_row = int(''.join([c for c in start_cell_str if c.isdigit()])) if any(c.isdigit() for c in start_cell_str) else 1
            
            # Estrai colonna e riga finale originali
            end_col_letters = ''.join([c for c in end_cell_str if not c.isdigit()])
            end_row_orig = int(''.join([c for c in end_cell_str if c.isdigit()])) if any(c.isdigit() for c in end_cell_str) else start_row
            
            # Converti le lettere delle colonne in indici
            try:
                start_col_idx = column_index_from_string(start_col_letters)
                end_col_idx = column_index_from_string(end_col_letters)
            except Exception:
                start_col_idx = 1
                end_col_idx = ws.max_column
            
            # Trova l'ultima riga con dati SOLO nelle colonne della tabella
            last_row_with_data = start_row
            # Partiamo dal basso verso l'alto per trovare l'ultima riga piena
            for row_idx in range(ws.max_row, start_row, -1):
                row_has_data = False
                for col_idx in range(start_col_idx, end_col_idx + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        last_row_with_data = row_idx
                        row_has_data = True
                        break
                if row_has_data:
                    break
            
            new_last_row = last_row_with_data
            new_end_cell = f"{end_col_letters}{new_last_row}"
            pandetta_table.ref = f"{start_cell_str}:{new_end_cell}"
            if hasattr(pandetta_table, 'autoFilter') and pandetta_table.autoFilter:
                pandetta_table.autoFilter.ref = pandetta_table.ref

    # Riordina le righe dati per N.RIF (crescente) prima di salvare
    # Questo risolve il problema delle nuove righe aggiunte in fondo al file
    rows_with_data = []
    for r_idx in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() for v in row_vals):
            id_val = ws.cell(row=r_idx, column=rif_col_idx).value
            if isinstance(id_val, float) and id_val.is_integer():
                id_val = int(id_val)
            rows_with_data.append((id_val, r_idx, row_vals))
    
    # Ordina per RIF (crescente), mettendo le righe senza RIF in fondo
    def sort_key(item):
        id_val = item[0]
        if id_val is None or str(id_val).strip() == '':
            return (1, 0)  # Metti righe senza RIF in fondo
        try:
            return (0, int(id_val))
        except (ValueError, TypeError):
            return (0, 0)
    
    rows_with_data.sort(key=sort_key)
    
    # Compila i valori delle righe ordinate
    header_vals = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    
    # Copia i dati riordinati
    for target_r_idx, (_, source_r_idx, row_vals) in enumerate(rows_with_data, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            ws.cell(row=target_r_idx, column=c_idx).value = val
    
    # Pulisci righe vuote in fondo
    for r_idx in range(len(rows_with_data) + 2, ws.max_row + 2):
        for c_idx in range(1, ws.max_column + 1):
            ws.cell(row=r_idx, column=c_idx).value = None

    wb.save(out_path)

if __name__ == '__main__':
    main()
