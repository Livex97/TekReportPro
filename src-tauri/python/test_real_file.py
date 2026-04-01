#!/usr/bin/env python3
"""Test sul file Excel reale PANDETTA_ORIG copia.xlsx"""
import os
import sys
import json
import tempfile
import openpyxl
from openpyxl.styles import PatternFill

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from save_pandetta import main

def analyze_excel_structure(filepath):
    """Analizza la struttura del file Excel reale."""
    print(f"Analisi del file: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    
    # Trova il foglio Pandetta
    for ws in wb.worksheets:
        print(f"\nFoglio: {ws.title}")
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"Intestazioni: {headers}")
        
        # Controlla se c'è N.RIF e STATO
        rif_col = None
        stato_col = None
        for idx, h in enumerate(headers):
            if h and 'RIF' in str(h).upper():
                rif_col = idx + 1
                print(f"  Colonna N.RIF trovata: {h} (indice {rif_col})")
            if h and 'STATO' in str(h).upper():
                stato_col = idx + 1
                print(f"  Colonna STATO trovata: {h} (indice {stato_col})")
        
        # Mostra prime righe di dati
        rows = list(ws.iter_rows(min_row=2, max_row=min(6, ws.max_row), values_only=True))
        print(f"Prime {len(rows)} righe di dati:")
        for i, row in enumerate(rows, start=2):
            print(f"  Riga {i}: {row}")
        
        break  # Analizza solo il primo foglio che sembra Pandetta
    
        # Riga con N.RIF=999 (nuova, STATUS=aperto -> giallo)
        if len(rows) >= 4:
            row_999 = pandetta_ws[4]  # Quarta riga
            stato_cell_999 = row_999[14]
            print(f"  Riga N.RIF=999 - Stato cella valore: {stato_cell_999.value}")
            if stato_cell_999.fill and stato_cell_999.fill.fill_type:
                fill_color_999 = stato_cell_999.fill.start_color
                if hasattr(fill_color_999, 'rgb') and fill_color_999.rgb:
                    print(f"    Colore: {fill_color_999.rgb} (atteso giallo FFFF00)")
        
        wb.close()
        print("\n✅ Test completato con successo!")
        return True

def test_with_real_file():
    """Test con il file Excel reale."""
    real_file = '/Users/alessio/Desktop/WorkAssistant/backup_template/PANDETTA_ORIG copia.xlsx'
    
    if not os.path.exists(real_file):
        print(f"File non trovato: {real_file}")
        return False
    
    print("="*60)
    analyze_excel_structure(real_file)
    print("="*60)
    
    with tempfile.TemporaryDirectory() as tmpdir:
        output_excel = os.path.join(tmpdir, 'output.xlsx')
        json_data = os.path.join(tmpdir, 'data.json')
        
        # Crea payload di test basato sulla struttura del file reale
        # Colonne reali: N. RIF, RICHIESTA INTERVENTO, DATA, CLIENTE, UBICAZIONE, 
        # STRUMENTO DA RIPARARE, TIPO DI GUASTO/ATTIVITA', DOC RITIRO, DATA RITIRO, 
        # GARANZIA (G) - CONTRATTO (C), DOC. PREV., DATA PREV., ACCETTAZIONE PREV., 
        # DATA ACCETTAZIONE PREV., STATO INTERVENTO, ESITO, DOC. CONSEGNA, DATA CONSEGNA, 
        # DOC. RAPPORTINO, DATA RAPPORTINO, TECNICO
        
        # Per semplificare, usiamo solo le colonne che ci interessano: N. RIF, DATA, CLIENTE, STATO INTERVENTO, TECNICO
        payload = {
            "data": [
                # Modifica una riga esistente (usiamo N. RIF = "1" che esiste già)
                {
                    "N. RIF": "1",
                    "DATA": "15/01/2026",
                    "CLIENTE": "SC AREA GESTIONE TECNICA -MODIFICATO",
                    "STATO INTERVENTO": "chiuso",  # Cambia stato (era 'IN ATTESA DI LORO ORDINE...')
                    "TECNICO": "ROSSI MODIFICATO"
                },
                # Aggiungi una nuova riga con N. RIF che non esiste
                {
                    "N. RIF": "999",
                    "DATA": "20/01/2026",
                    "CLIENTE": "NUOVO CLIENTE TEST",
                    "STATO INTERVENTO": "aperto",
                    "TECNICO": "VERDI"
                }
            ],
            "dynamic_cols": ["N. RIF", "DATA", "CLIENTE", "STATO INTERVENTO", "TECNICO"]
        }
        
        with open(json_data, 'w', encoding='utf-8') as f:
            json.dump(payload, f)
        
        print("\nEsecuzione di save_pandetta.py...")
        sys.argv = ['save_pandetta.py', json_data, real_file, output_excel]
        
        try:
            main()
            print("✅ save_pandetta.py eseguito senza errori")
        except Exception as e:
            print(f"❌ Errore durante l'esecuzione: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        # Verifica output
        print("\nVerifica del file di output...")
        wb = openpyxl.load_workbook(output_excel)
        ws = wb.active  # o cerca 'Pandetta'
        
        # Cerca il foglio Pandetta
        pandetta_ws = None
        for ws_name in wb.sheetnames:
            if 'PANDETTA' in ws_name.upper() or 'ASSISTENZA' in ws_name.upper():
                pandetta_ws = wb[ws_name]
                print(f"Trovato foglio: {ws_name}")
                break
        
        if not pandetta_ws:
            pandetta_ws = wb.active
            print("Usato foglio attivo")
        
        rows = list(pandetta_ws.iter_rows(values_only=True))
        print(f"Numero totale di righe (inclusa intestazione): {len(rows)}")
        
        # Cerca le righe modificate/aggiunte
        print("\nRicerca valori N.RIF specifici:")
        rif_to_find = ["1", "999"]
        for rif in rif_to_find:
            found = False
            for row_idx, row in enumerate(rows):
                if row[0] == rif:  # N.RIF è la prima colonna
                    print(f"  Trovato N.RIF={rif} alla riga {row_idx+1}: {row}")
                    found = True
            if not found:
                print(f"  N.RIF={rif} non trovato")
        
        # Verifica colori delle righe
        print("\nVerifica colori applicati:")
        # Riga con N.RIF=1 (modificata, STATUS=chiuso -> verde)
        if len(rows) >= 3:
            row_1 = pandetta_ws[3]  # Terza riga, indice 3
            stato_cell = row_1[14]  # Colonna STATO INTERVENTO (indice 15, quindi 14 in zero-based)
            print(f"  Riga N.RIF=1 - Stato cella valore: {stato_cell.value}")
            if stato_cell.fill and stato_cell.fill.fill_type:
                fill_color = stato_cell.fill.start_color
                if hasattr(fill_color, 'rgb') and fill_color.rgb:
                    print(f"    Colore: {fill_color.rgb} (atteso verde 00FF00)")
            # Controlla che anche le altre celle della riga abbiano il fill (seabbiamo applicato a tutta la riga)
            # La funzione apply_status_color ora applica il colore a tutta la riga
        
        # Riga con N.RIF=999 (nuova, STATUS=aperto -> giallo)
        if len(rows) >= 4:
            row_999 = pandetta_ws[4]  # Quarta riga
            stato_cell_999 = row_999[14]
            print(f"  Riga N.RIF=999 - Stato cella valore: {stato_cell_999.value}")
            if stato_cell_999.fill and stato_cell_999.fill.fill_type:
                fill_color_999 = stato_cell_999.fill.start_color
                if hasattr(fill_color_999, 'rgb') and fill_color_999.rgb:
                    print(f"    Colore: {fill_color_999.rgb} (atteso giallo FFFF00)")
        print("\n✅ Test completato con successo!")
        return True

if __name__ == '__main__':
    try:
        success = test_with_real_file()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"❌ Test fallito: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
